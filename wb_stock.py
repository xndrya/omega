import openpyxl
import pandas as pd

"""
скрипт для записи наименований позиций в эксель-файл дефицита для вайлдберриз
"""

# путь до файла
wb = openpyxl.load_workbook('input/Дефицит.xlsx')
# путь до справочника
directory = pd.read_excel('data/directory.xlsx', sheet_name='directory', index_col="Штрихкод")

sheet = wb["Заказ"]
num = 0
# вычисляем кол-во позиций
for k in range(2, 150, 1):
    if sheet.cell(row=k, column=1).value is not None:
        num += 1
    else:
        break

for i in range(2, 2 + num, 1):
    item = int(sheet.cell(row=i, column=14).value)
    sheet.cell(row=i, column=6).value = directory["Название"][item]

wb.save('output/Дефицит.xlsx')
