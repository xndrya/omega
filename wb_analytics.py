import openpyxl
import pandas as pd

"""
скрипт для записи наименований позиций в эксель-файл с остатками для вайлдберриз
"""

# путь до счета
wb = openpyxl.load_workbook('input/SupplierStock.xlsx')
# путь до справочника
directory = pd.read_excel('data/directory.xlsx', sheet_name='directory', index_col="Штрихкод")

sheet = wb["SupplierStock"]
num = 0
# вычисляем кол-во позиций
for k in range(3, 150, 1):
    if sheet.cell(row=k, column=1).value is not None:
        num += 1
    else:
        break

for i in range(3, 3 + num, 1):
    item = int(sheet.cell(row=i, column=5).value)
    sheet.cell(row=i, column=2).value = directory["Название"][item]

wb.save('output/SupplierStock.xlsx')
