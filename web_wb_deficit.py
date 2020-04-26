import openpyxl
import pandas as pd


def make_file(file_path, type):
    """
    Функция для добавления наименований позиций в файлы с дефицитом, остатками, ценами
    :param Дефицит/Остатки/Цены:
    """
    # путь до файла
    wb = openpyxl.load_workbook(file_path)
    # путь до справочника
    directory = pd.read_excel('D:\Projects\WB scripts\data\directory.xlsx', sheet_name='directory',
                              index_col="Штрихкод")

    params = {
        "Дефицит": {
            "sheet": "Заказ",
            "k_start": 2,
            "item_column": 14,
            "sheet_cell_column": 6
        },
        "Остатки": {
            "sheet": "SupplierStock",
            "k_start": 3,
            "item_column": 5,
            "sheet_cell_column": 2
        },
        "Цены": {
            "sheet": "Общий отчет",
            "k_start": 2,
            "item_column": 6,
            "sheet_cell_column": 19
        }
    }

    sheet = wb[params[type]["sheet"]]
    num = 0
    # вычисляем кол-во позиций
    for k in range(params[type]["k_start"], 150, 1):
        if sheet.cell(row=k, column=1).value is not None:
            num += 1
        else:
            break

    for i in range(2, 2 + num, 1):
        item = int(sheet.cell(row=i, column=params[type]["item_column"]).value)
        sheet.cell(row=i, column=params[type]["sheet_cell_column"]).value = directory["Название"][item]

    wb.save(file_path)


if __name__ == "__main__":
    make_file("input/Дефицит.xlsx", "Дефицит")
