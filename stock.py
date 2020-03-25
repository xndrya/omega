import pandas as pd
import openpyxl

"""
скрипт для записи наименований позиций в эксель-файл  дефицита  для вайлдберриз
"""

# путь до счета
wb = openpyxl.load_workbook('input/SupplierStock.xlsx')
# путь до справочника
dir = pd.read_excel('data/directory.xlsx', sheet_name='directory',
                    index_col="Штрихкод")

sheet = wb["SupplierStock"]
num = 0
# вычисляем кол-во позиций
for k in range(2, 150, 1):
    if sheet.cell(row=k, column=1).value != None:
        num += 1
    else:
        break

for i in range(2, 2 + num, 1):
    item = int(sheet.cell(row=i, column=5).value)
    sheet.cell(row=i, column=6).value = dir["Название"][item]

wb.save('output/SupplierStock.xlsx')
