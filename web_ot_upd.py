# -*- coding: cp1251 -*-
import datetime
import openpyxl
import pandas as pd
from lxml import etree


def make_ot_upd(file_path, type):
    # путь до счета
    invoice = openpyxl.load_workbook(file_path, read_only=True)
    sheet_invoice = invoice["TDSheet"]

    # путь до справочника
    dir = pd.read_excel('D:\Projects\WB scripts\data\directory.xlsx', sheet_name='directory',
                        index_col="Название")

    # путь до файла xml
    if type == "OT":
        template_path = 'D:\\Projects\\WB scripts\\data\\template_main3.xml'
    elif type == "WB":
        template_path = 'D:\\Projects\\WB scripts\\data\\template_main2.xml'
    else:
        template_path = 'D:\\Projects\\WB scripts\\data\\template_main.xml'

    doc = etree.parse(template_path)
    num = 0
    # вычисляем кол-во позиций в счете с 25 по 150 ячейку
    for k in range(25, 150, 1):
        if sheet_invoice.cell(row=k, column=2).value != None:
            num = sheet_invoice.cell(row=k, column=2).value
        else:
            break

    total_sum_without_nds = 0
    total_sum = 0
    total_quantity = 0
    table_root = doc.find("//ТаблСчФакт")
    i = 0

    for i in range(num):
        # количество
        quantity = sheet_invoice.cell(row=i + 25, column=20).value

        # ниже получаем размер ндс, сумму по позиции и вычисляем цену без ндс для нее
        price = sheet_invoice.cell(row=i + 25, column=25).value
        nds = sheet_invoice.cell(row=i + 25, column=29).value
        price_without_nds = round(price / (1 + nds * 0.01), 2)
        total_sum_per_good = sheet_invoice.cell(row=i + 25, column=37).value

        # создаем элемент СведТов и начинаем наполнять
        good = etree.Element("СведТов")
        good.set("КолТов", str(quantity) + ".000")

        # наименование товара берем из счета
        item = sheet_invoice.cell(row=i + 25, column=4).value
        good.set("НаимТов", item)
        good.set("НалСт", str(nds) + "%")
        good.set("НомСтр", str(i + 1))
        good.set("ОКЕИ_Тов", "796")
        sum_price_without_nds = round(price_without_nds * quantity, 2)
        good.set("СтТовБезНДС", str(sum_price_without_nds))
        good.set("СтТовУчНал", str(total_sum_per_good))
        good.set("ЦенаТов", str(price_without_nds))

        acc = etree.SubElement(good, 'Акциз')
        acc_free = etree.SubElement(acc, 'БезАкциз')
        acc_free.text = "без акциза"

        tax_sum = etree.SubElement(good, 'СумНал')
        tax_sum2 = etree.SubElement(tax_sum, 'СумНал')
        item_nds = round(total_sum_per_good - (price_without_nds * quantity), 2)
        tax_sum2.text = str(item_nds)

        td = etree.SubElement(good, 'СвТД')
        td.set("КодПроисх", str(dir["Код страны"][item]))
        td.set("НомерТД", str(dir["ГТД"][item]))

        addit = etree.SubElement(good, 'ДопСведТов')
        addit.set("КодТов", str(dir["Штрихкод"][item]))
        addit.set("КрНаимСтрПр", dir["Страна"][item])
        addit.set("НаимЕдИзм", "шт")
        addit.set("ПрТовРаб", "1")

        info = etree.SubElement(good, 'ИнфПолФХЖ2')
        info.set("Значен", str(dir["Штрихкод"][item]))
        info.set("Идентиф", "ШК")

        info2 = etree.SubElement(good, 'ИнфПолФХЖ2')
        info2.set("Значен", str(dir["Штрихкод"][item]))
        info2.set("Идентиф", "Код")

        info3 = etree.SubElement(good, 'ИнфПолФХЖ2')
        info3.set("Значен", str(dir["Штрихкод"][item]))
        info3.set("Идентиф", "Ид")

        info4 = etree.SubElement(good, 'ИнфПолФХЖ2')
        info4.set("Значен", str(dir["Штрихкод"][item]))
        info4.set("Идентиф", "КодПоставщика")

        info5 = etree.SubElement(good, 'ИнфПолФХЖ2')
        info5.set("Значен", item)
        info5.set("Идентиф", "НазваниеПокупателя")

        info6 = etree.SubElement(good, 'ИнфПолФХЖ2')
        info6.set("Значен", item)
        info6.set("Идентиф", "НазваниеПоставщика")

        info7 = etree.SubElement(good, 'ИнфПолФХЖ2')
        info7.set("Значен", str(dir["Артикул"][item]))
        info7.set("Идентиф", "ИД")

        table_root = doc.find("//ТаблСчФакт")

        # считаем итоговые суммы с ндс и без и количество
        total_sum_without_nds += sum_price_without_nds
        total_sum += total_sum_per_good
        total_quantity += quantity

        # вставляем элемент
        table_root.insert(i, good)

    # проставляем итоговые суммы с ндс и без и количество
    total = etree.Element("ВсегоОпл")
    total.set("СтТовБезНДСВсего", str(round(total_sum_without_nds, 2)))
    total.set("СтТовУчНалВсего", str(total_sum))
    total_nds = etree.SubElement(total, "СумНалВсего")
    total_nds2 = etree.SubElement(total_nds, "СумНал")
    total_nds3 = round(total_sum - total_sum_without_nds, 2)
    total_nds2.text = str(total_nds3)
    total_quantity2 = etree.SubElement(total, "КолНеттоВс")
    total_quantity2.text = str(total_quantity)
    table_root.insert(i + 1, total)

    # вычисляем текущую  дату, время и проставляем в документ
    date = datetime.datetime.today().strftime("%d.%m.%Y")
    time = datetime.datetime.today().strftime("%H.%M.%S")

    doc.find("//Документ").attrib['ДатаИнфПр'] = date
    doc.find("//Документ").attrib['ВремИнфПр'] = time
    doc.find("//СвСчФакт").attrib['ДатаСчФ'] = date
    doc.find("//СвПер").attrib['ДатаПер'] = date
    doc.findall("//ТекстИнф")[0].attrib['Значен'] = datetime.datetime.today().strftime("%d.%m.%Y %H:%M:%S")
    doc.findall("//ТекстИнф")[2].attrib['Значен'] = date
    doc.findall("//ТекстИнф")[4].attrib['Значен'] = date
    doc.findall("//ТекстИнф")[5].attrib['Значен'] = date

    f = doc.write("D:\\Projects\\WB scripts\\output\\output.xml", encoding="cp1251", pretty_print=True,
                  xml_declaration=True)
    return f


if __name__ == "__main__":
    make_ot_upd('input/invoice.xlsx')
