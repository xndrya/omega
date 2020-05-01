import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font


def make_ot_invoice(file_path):
    # путь до счета
    wb = openpyxl.load_workbook(file_path)
    # путь до справочника
    dir = pd.read_excel('D:\Projects\WB scripts\data\directory.xlsx', sheet_name='directory', index_col="Название")
    # наименование листа с товарами в счете
    sheet = wb["TDSheet"]

    # форматируем ячейки
    sheet.merge_cells('AS23:AS24')
    sheet.column_dimensions['AS'].width = 15
    sheet['AS23'].value = "Штрихкод"
    sheet['AS23'].font = Font(name='Arial', size=10, bold=True)
    sheet['AS23'].alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('AT23:AT24')
    sheet['AT23'].value = "Страна"
    sheet['AT23'].font = Font(name='Arial', size=10, bold=True)
    sheet['AT23'].alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('AU23:AU24')
    sheet.column_dimensions['AU'].width = 30
    sheet['AU23'].value = "ГТД"
    sheet['AU23'].font = Font(name='Arial', size=10, bold=True)
    sheet['AU23'].alignment = Alignment(horizontal='center', vertical='center')

    # вычисляем кол-во позиций
    for k in range(25, 150, 1):
        if sheet.cell(row=k, column=2).value != None:
            num = sheet.cell(row=k, column=2).value
        else:
            break

    for i in range(25, 25 + num, 1):
        sheet.cell(row=i, column=45).value = dir["Штрихкод"][sheet.cell(row=i, column=4).value]
        sheet.cell(row=i, column=46).value = dir["Страна"][sheet.cell(row=i, column=4).value]
        sheet.cell(row=i, column=47).value = dir["ГТД"][sheet.cell(row=i, column=4).value]

    wb.save(file_path)


if __name__ == "__main__":
    make_ot_invoice('C:\\Users\\X\\Desktop\\Сломан.xlsx')
